"""Admin database Excel export routes."""

import io
import zipfile
from datetime import datetime
from pathlib import Path
from flask import current_app, render_template, request, flash, redirect, url_for, send_file, jsonify
from auth.roles import roles_required
from utils.env_loader import get_env
from utils.log_sanitize import sanitize_log_value
from utils.sensitive_operations import requires_reauth, log_export_initiated, log_export_completed, log_export_failed
from db_transaction_manager import get_db_session
from sqlalchemy import text
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.filename_utils import sanitize_export_filename


@requires_reauth("database_excel_export")
@roles_required("admin")
def database_excel_export():
    """Handle database Excel export functionality."""
    if request.method == "POST":
        # Log export initiation
        log_export_initiated("database_excel_export", {'tables': request.form.getlist("tables")})
        
        try:
            # Get selected tables from form
            selected_tables = request.form.getlist("tables")
            
            if not selected_tables:
                flash("Please select at least one table to export.", "danger")
                return redirect(url_for("admin.database_excel_export"))
            
            # Create a ZIP file containing all Excel files
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            zip_filename = f"database_export_{timestamp}.zip"
            
            # Create a BytesIO buffer for the ZIP file
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for table_name in selected_tables:
                    try:
                        # Export table to Excel
                        excel_data = _export_table_to_excel(table_name)
                        if excel_data:
                            # Add Excel file to ZIP
                            excel_filename = sanitize_export_filename(table_name, "xlsx")
                            zip_file.writestr(excel_filename, excel_data)
                            current_app.logger.info(
                                "Added %s to export",
                                sanitize_log_value(excel_filename),
                            )
                        else:
                            current_app.logger.warning(
                                "No data exported for table: %s",
                                sanitize_log_value(table_name),
                            )
                    except Exception as e:
                        current_app.logger.error(
                            "Error exporting table %s: %s",
                            sanitize_log_value(table_name),
                            sanitize_log_value(e),
                        )
                        flash(f"Error exporting table {table_name}. Please check logs.", "warning")
            
            # Reset buffer position
            zip_buffer.seek(0)
            
            # Calculate total row count from all tables
            total_rows = len(selected_tables)  # Approximate - one per table
            
            # Save to temp file for hash calculation
            import tempfile
            temp_file = Path(tempfile.gettempdir()) / zip_filename
            with open(temp_file, 'wb') as f:
                f.write(zip_buffer.getvalue())
            
            # Log successful export
            log_export_completed("database_excel_export", str(temp_file), total_rows)
            
            # Log the export operation
            current_app.logger.info(
                "Database Excel export created: %s with %s tables",
                sanitize_log_value(zip_filename),
                sanitize_log_value(len(selected_tables)),
            )
            
            # Reset buffer for sending
            zip_buffer.seek(0)
            
            # Send ZIP file to user
            response = send_file(
                zip_buffer,
                as_attachment=True,
                download_name=zip_filename,
                mimetype='application/zip'
            )
            
            # Clean up temp file
            try:
                temp_file.unlink()
            except Exception:
                pass
            
            return response
            
        except Exception as e:
            current_app.logger.error(
                "Error creating database Excel export: %s",
                sanitize_log_value(e),
            )
            flash("Error creating database Excel export. Please check logs.", "danger")
            log_export_failed("database_excel_export", str(e))
    
    # GET request - show the export page
    return render_template("admin/database_excel_export.html")


def _export_table_to_excel(table_name):
    """Export a single table to Excel format."""
    try:
        with get_db_session() as db:
            # Get table data, excluding sensitive columns
            # Validate table_name to prevent SQL injection
            if get_env("DATABASE_URL", "").startswith("postgresql://"):
                check_query = text("SELECT 1 FROM pg_tables WHERE schemaname = 'public' AND tablename = :t")
            else:
                check_query = text("SELECT 1 FROM sqlite_master WHERE type='table' AND name = :t")
            
            if not db.execute(check_query, {"t": table_name}).fetchone():
                current_app.logger.warning(
                    "Attempted export of invalid table: %s",
                    sanitize_log_value(table_name),
                )
                return None

            # Get table data, excluding sensitive columns
            if table_name == 'users':
                # Exclude password_hash from users table
                query = text("""
                    SELECT id, username, is_active, is_locked_until, full_name, phone, designation,
                           email, year_of_joining, last_date_of_service, created_at, updated_at,
                           file_upload_quota, file_upload_count, timezone
                    FROM users
                """)
            else:
                # Safe because table_name is validated against allowlist above
                query = text(f"SELECT * FROM {table_name}")
            
            result = db.execute(query)
            rows = result.fetchall()
            
            if not rows:
                return None
            
            # Convert to DataFrame
            columns = list(result.keys())
            df = pd.DataFrame(rows, columns=columns)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # Add headers (start from row 1)
            row = 1
            for col, header in enumerate(columns, start=1):
                cell = ws.cell(row=row, column=col, value=header.replace('_', ' ').title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Add data
            row += 1
            for _, data_row in df.iterrows():
                for col, value in enumerate(data_row, start=1):
                    # Handle datetime objects and NaN values
                    if hasattr(value, 'strftime'):
                        try:
                            value = value.strftime("%Y-%m-%d %H:%M:%S")
                        except (ValueError, AttributeError):
                            value = "N/A"
                    elif value is None or (isinstance(value, float) and (value != value)):  # Check for NaN
                        value = "N/A"
                    else:
                        value = str(value)
                    
                    ws.cell(row=row, column=col, value=value)
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                
                row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, ValueError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create a BytesIO buffer to save workbook
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
    except Exception as e:
        current_app.logger.error(
            "Error exporting table %s to Excel: %s",
            sanitize_log_value(table_name),
            sanitize_log_value(e),
        )
        return None


@roles_required("admin")
def get_database_tables():
    """Get list of all database tables as JSON for AJAX requests."""
    try:
        with get_db_session() as db:
            # Get all table names
            if get_env("DATABASE_URL", "").startswith("postgresql://"):
                result = db.execute(text("""
                    SELECT tablename FROM pg_tables
                    WHERE schemaname = 'public'
                    ORDER BY tablename
                """))
            else:  # SQLite
                result = db.execute(text("""
                    SELECT name FROM sqlite_master
                    WHERE type='table'
                    ORDER BY name
                """))
            
            tables = [row[0] for row in result]
            
            # Get row counts for each table
            table_info = []
            for table in tables:
                try:
                    count_result = db.execute(text(f"SELECT COUNT(*) FROM {table}"))
                    row_count = count_result.scalar()
                    table_info.append({
                        "name": table,
                        "row_count": row_count
                    })
                except Exception as e:
                    current_app.logger.warning(
                        "Could not get row count for %s: %s",
                        sanitize_log_value(table),
                        sanitize_log_value(e),
                    )
                    table_info.append({
                        "name": table,
                        "row_count": 0
                    })
            
            return jsonify({
                "tables": table_info,
                "total_tables": len(table_info)
            })
            
    except Exception as e:
        current_app.logger.error(
            "Error getting database tables: %s",
            sanitize_log_value(e),
        )
        return jsonify({"error": "Internal server error"}), 500
