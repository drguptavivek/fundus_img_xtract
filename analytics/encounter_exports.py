"""Encounter export builders for analytics downloads."""

from __future__ import annotations

import io
import json
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import sqlalchemy as sa
from sqlalchemy.orm import Session as SASession, selectinload

from models import (
    DiabeticRetinopathyReport,
    EncounterFile,
    EncounterSetImage,
    GlaucomaResultsCleaned,
    LabUnit,
    PatientEncounters,
    User,
)
from utils.hospital_scoping import apply_scoping


@dataclass(frozen=True)
class EncounterExportFilters:
    hospital_id: int | None = None
    lab_unit_id: int | None = None
    capture_date: date | None = None
    start_date: date | None = None
    end_date: date | None = None
    project_ids: tuple[int, ...] = ()
    include_classical: bool | None = None


_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _cell_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        return json.dumps(value, sort_keys=True)
    if value is None:
        return ""
    return value


def _write_sheet(ws, headers: list[str], rows: Iterable[dict[str, Any]]) -> None:
    if getattr(ws.parent, "write_only", False):
        ws.append(headers)
        for row_data in rows:
            ws.append([_cell_value(row_data.get(header)) for header in headers])
        return

    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for row_data in rows:
        ws.append([_cell_value(row_data.get(header)) for header in headers])
        for cell in ws[ws.max_row]:
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in column), default=0)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, 12), 60)
    ws.freeze_panes = "A2"


def _workbook_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _scoped_lab_unit_ids(db: SASession, user: User, filters: EncounterExportFilters) -> list[int]:
    query = db.query(LabUnit.id)
    query = apply_scoping(query, LabUnit, user, "analytics")
    if filters.hospital_id is not None:
        query = query.filter(LabUnit.hospital_id == filters.hospital_id)
    if filters.lab_unit_id is not None:
        query = query.filter(LabUnit.id == filters.lab_unit_id)
    return [int(row[0]) for row in query.all()]


def encounter_ids(
    db: SASession,
    user: User,
    filters: EncounterExportFilters,
) -> list[int]:
    """Return all scoped encounter IDs matching export filters."""

    lab_unit_ids = _scoped_lab_unit_ids(db, user, filters)
    if not lab_unit_ids:
        return []

    query = db.query(PatientEncounters.id).filter(PatientEncounters.lab_unit_id.in_(lab_unit_ids))
    if filters.capture_date is not None:
        query = query.filter(PatientEncounters.capture_date_dt == filters.capture_date)
    if filters.start_date is not None:
        query = query.filter(PatientEncounters.capture_date_dt >= filters.start_date)
    if filters.end_date is not None:
        query = query.filter(PatientEncounters.capture_date_dt <= filters.end_date)
    if filters.project_ids or filters.include_classical is not None:
        project_conditions = []
        if filters.project_ids:
            project_conditions.append(PatientEncounters.project_id.in_(filters.project_ids))
        if filters.include_classical:
            project_conditions.append(PatientEncounters.project_id.is_(None))
        if project_conditions:
            query = query.filter(sa.or_(*project_conditions))
        else:
            query = query.filter(PatientEncounters.id.is_(None))

    return [
        int(row[0])
        for row in query.order_by(PatientEncounters.capture_date_dt.desc().nullslast(), PatientEncounters.id.desc()).all()
    ]


def _load_encounters(
    db: SASession,
    user: User,
    filters: EncounterExportFilters,
) -> list[PatientEncounters]:
    ids = encounter_ids(db, user, filters)
    if not ids:
        return []
    encounters = (
        db.query(PatientEncounters)
        .filter(PatientEncounters.id.in_(ids))
        .options(
            selectinload(PatientEncounters.lab_unit).selectinload(LabUnit.hospital),
            selectinload(PatientEncounters.project),
            selectinload(PatientEncounters.dr_reports),
            selectinload(PatientEncounters.glaucoma_reports),
            selectinload(PatientEncounters.glaucoma_results_cleaned),
            selectinload(PatientEncounters.encounter_files).selectinload(EncounterFile.camera),
            selectinload(PatientEncounters.encounter_set_images).selectinload(EncounterSetImage.camera),
            selectinload(PatientEncounters.zip_file),
        )
        .all()
    )
    order_map = {encounter_id: index for index, encounter_id in enumerate(ids)}
    encounters.sort(key=lambda encounter: order_map.get(encounter.id, 10**9))
    return encounters


def _latest_dr_report(encounter: PatientEncounters) -> DiabeticRetinopathyReport | None:
    if not encounter.dr_reports:
        return None
    return max(encounter.dr_reports, key=lambda report: report.id)


def _latest_glaucoma_cleaned(encounter: PatientEncounters) -> GlaucomaResultsCleaned | None:
    if not encounter.glaucoma_results_cleaned:
        return None
    return max(
        encounter.glaucoma_results_cleaned,
        key=lambda row: row.updated_at or row.created_at or datetime.min,
    )


def _encounter_row(encounter: PatientEncounters) -> dict[str, Any]:
    lab_unit = encounter.lab_unit
    hospital = lab_unit.hospital if lab_unit else None
    latest_dr = _latest_dr_report(encounter)
    latest_glaucoma = _latest_glaucoma_cleaned(encounter)
    project = encounter.project
    return {
        "encounter_id": encounter.id,
        "encounter_uuid": encounter.uuid,
        "patient_name": encounter.name,
        "patient_id": encounter.patient_id,
        "project_id": project.id if project else "",
        "project_code": project.code if project else "",
        "project_title": project.title if project else "",
        "is_classical_encounter": encounter.project_id is None,
        "capture_date_text": encounter.capture_date,
        "capture_date": encounter.capture_date_dt,
        "hospital_id": hospital.id if hospital else "",
        "hospital_name": hospital.name if hospital else "",
        "lab_unit_id": lab_unit.id if lab_unit else "",
        "lab_unit_name": lab_unit.name if lab_unit else "",
        "zip_file_id": encounter.zip_file_id,
        "zip_filename": encounter.zip_file.zip_filename if encounter.zip_file else "",
        "encounter_verified_status": encounter.encounter_verified_status,
        "encounter_verified_by": encounter.encounter_verified_by,
        "encounter_verified_at": encounter.encounter_verified_at,
        "dr_verified_status": encounter.dr_verified_status,
        "dr_verified_by": encounter.dr_verified_by,
        "dr_verified_at": encounter.dr_verified_at,
        "glaucoma_verified_status": encounter.glaucoma_verified_status,
        "glaucoma_verified_by": encounter.glaucoma_verified_by,
        "glaucoma_verified_at": encounter.glaucoma_verified_at,
        "is_set_based": encounter.is_set_based,
        "classical_image_count": len(encounter.encounter_files),
        "encounter_set_image_count": len(encounter.encounter_set_images),
        "image_count": len(encounter.encounter_files) + len(encounter.encounter_set_images),
        "latest_dr_result": latest_dr.result if latest_dr else "",
        "latest_dr_qualitative_result": latest_dr.qualitative_result if latest_dr else "",
        "latest_dr_report_file_name": latest_dr.report_file_name if latest_dr else "",
        "latest_glaucoma_result": latest_glaucoma.result if latest_glaucoma else "",
        "latest_glaucoma_qualitative_result": latest_glaucoma.qualitative_result if latest_glaucoma else "",
        "latest_glaucoma_vcdr_right": latest_glaucoma.vcdr_right_num if latest_glaucoma else "",
        "latest_glaucoma_vcdr_left": latest_glaucoma.vcdr_left_num if latest_glaucoma else "",
        "latest_glaucoma_report_uuid": latest_glaucoma.report_uuid if latest_glaucoma else "",
        "latest_glaucoma_report_file_name": latest_glaucoma.report_file_name if latest_glaucoma else "",
        "dr_ocr_count": len(encounter.dr_reports),
        "glaucoma_ocr_count": len(encounter.glaucoma_results_cleaned),
        "remarks": encounter.remarks,
        "metadata_json": encounter.metadata_json,
    }


def _flatten_ocr_columns(encounter: PatientEncounters, row: dict[str, Any], max_dr: int, max_glaucoma: int) -> None:
    dr_reports = sorted(encounter.dr_reports, key=lambda item: item.id)
    for index in range(max_dr):
        prefix = f"dr_ocr_{index + 1}"
        report = dr_reports[index] if index < len(dr_reports) else None
        row[f"{prefix}_report_id"] = report.id if report else ""
        row[f"{prefix}_report_uuid"] = report.uuid if report else ""
        row[f"{prefix}_result"] = report.result if report else ""
        row[f"{prefix}_qualitative_result"] = report.qualitative_result if report else ""
        row[f"{prefix}_report_file_name"] = report.report_file_name if report else ""

    glaucoma_rows = sorted(encounter.glaucoma_results_cleaned, key=lambda item: item.id)
    for index in range(max_glaucoma):
        prefix = f"glaucoma_ocr_{index + 1}"
        glaucoma = glaucoma_rows[index] if index < len(glaucoma_rows) else None
        row[f"{prefix}_cleaned_id"] = glaucoma.id if glaucoma else ""
        row[f"{prefix}_report_uuid"] = glaucoma.report_uuid if glaucoma else ""
        row[f"{prefix}_result"] = glaucoma.result if glaucoma else ""
        row[f"{prefix}_qualitative_result"] = glaucoma.qualitative_result if glaucoma else ""
        row[f"{prefix}_vcdr_right"] = glaucoma.vcdr_right_num if glaucoma else ""
        row[f"{prefix}_vcdr_left"] = glaucoma.vcdr_left_num if glaucoma else ""
        row[f"{prefix}_original_vcdr_right"] = glaucoma.original_vcdr_right if glaucoma else ""
        row[f"{prefix}_original_vcdr_left"] = glaucoma.original_vcdr_left if glaucoma else ""
        row[f"{prefix}_report_file_name"] = glaucoma.report_file_name if glaucoma else ""
        row[f"{prefix}_created_at"] = glaucoma.created_at if glaucoma else ""
        row[f"{prefix}_updated_at"] = glaucoma.updated_at if glaucoma else ""


def export_encounter_data_xlsx(
    db: SASession,
    user: User,
    filters: EncounterExportFilters,
) -> bytes:
    """Build one flat encounter-level sheet with OCR report columns."""

    encounters = _load_encounters(db, user, filters)
    max_dr = max((len(encounter.dr_reports) for encounter in encounters), default=0)
    max_glaucoma = max((len(encounter.glaucoma_results_cleaned) for encounter in encounters), default=0)
    wb = Workbook()
    ws = wb.active
    ws.title = "Encounter OCR Data"

    encounter_headers = [
        "encounter_id",
        "encounter_uuid",
        "patient_name",
        "patient_id",
        "project_id",
        "project_code",
        "project_title",
        "is_classical_encounter",
        "capture_date_text",
        "capture_date",
        "hospital_id",
        "hospital_name",
        "lab_unit_id",
        "lab_unit_name",
        "zip_file_id",
        "zip_filename",
        "encounter_verified_status",
        "encounter_verified_by",
        "encounter_verified_at",
        "dr_verified_status",
        "dr_verified_by",
        "dr_verified_at",
        "glaucoma_verified_status",
        "glaucoma_verified_by",
        "glaucoma_verified_at",
        "is_set_based",
        "classical_image_count",
        "encounter_set_image_count",
        "image_count",
        "latest_dr_result",
        "latest_dr_qualitative_result",
        "latest_dr_report_file_name",
        "latest_glaucoma_result",
        "latest_glaucoma_qualitative_result",
        "latest_glaucoma_vcdr_right",
        "latest_glaucoma_vcdr_left",
        "latest_glaucoma_report_uuid",
        "latest_glaucoma_report_file_name",
        "dr_ocr_count",
        "glaucoma_ocr_count",
        "remarks",
        "metadata_json",
    ]
    for index in range(max_dr):
        prefix = f"dr_ocr_{index + 1}"
        encounter_headers.extend(
            [
                f"{prefix}_report_id",
                f"{prefix}_report_uuid",
                f"{prefix}_result",
                f"{prefix}_qualitative_result",
                f"{prefix}_report_file_name",
            ]
        )
    for index in range(max_glaucoma):
        prefix = f"glaucoma_ocr_{index + 1}"
        encounter_headers.extend(
            [
                f"{prefix}_cleaned_id",
                f"{prefix}_report_uuid",
                f"{prefix}_result",
                f"{prefix}_qualitative_result",
                f"{prefix}_vcdr_right",
                f"{prefix}_vcdr_left",
                f"{prefix}_original_vcdr_right",
                f"{prefix}_original_vcdr_left",
                f"{prefix}_report_file_name",
                f"{prefix}_created_at",
                f"{prefix}_updated_at",
            ]
        )

    rows = []
    for encounter in encounters:
        row = _encounter_row(encounter)
        _flatten_ocr_columns(encounter, row, max_dr, max_glaucoma)
        rows.append(row)
    _write_sheet(ws, encounter_headers, rows)

    return _workbook_bytes(wb)


_TASK_RESULT_HEADERS = [
    "source_view",
    "image_source",
    "image_id",
    "image_uuid",
    "filename",
    "eye_side",
    "patient_encounter_id",
    "patient_encounter_name",
    "patient_identifier",
    "project_id",
    "project_code",
    "project_title",
    "is_classical_encounter",
    "capture_date",
    "hospital_name",
    "lab_unit_name",
    "camera_name",
    "disease_name",
    "is_mydriatic",
    "is_pregraded",
    "task_id",
    "task_uuid",
    "task_state",
    "task_created_at",
    "resident_grade_id",
    "resident_grade",
    "resident_grader",
    "resident_grade_time",
    "resident_comment",
    "resident_features",
    "resident2_grade_id",
    "resident2_grade",
    "resident2_grader",
    "resident2_grade_time",
    "resident2_comment",
    "resident2_features",
    "arbitrator_grade_id",
    "arbitrator_grade",
    "arbitrator_grader",
    "arbitrator_grade_time",
    "arbitrator_comment",
    "arbitrator_features",
    "review_grade_id",
    "review_grade",
    "reviewer_name",
    "review_grade_time",
    "review_comment",
    "review_features",
    "regrade_adj_grade_ids",
    "regrade_adj_grades",
    "regrade_adj_graders",
    "regrade_adj_grade_times",
    "regrade_adj_comments",
    "regrade_adj_features",
    "aimodel_1_grade_id",
    "aimodel_1_grade",
    "aimodel_1_name",
    "aimodel_1_time",
    "aimodel_1_features",
    "aimodel_2_grade_id",
    "aimodel_2_grade",
    "aimodel_2_name",
    "aimodel_2_time",
    "aimodel_2_features",
    "aimodel_3_grade_id",
    "aimodel_3_grade",
    "aimodel_3_name",
    "aimodel_3_time",
    "aimodel_3_features",
    "consensus_grade",
    "consensus_method",
    "consensus_decider",
    "consensus_time",
    "last_updated",
]

_REGRADE_ADJ_HEADERS = {
    "regrade_adj_grade_ids",
    "regrade_adj_grades",
    "regrade_adj_graders",
    "regrade_adj_grade_times",
    "regrade_adj_comments",
    "regrade_adj_features",
}

_JOINED_TASK_HEADERS = _REGRADE_ADJ_HEADERS | {
    "project_id",
    "project_code",
    "project_title",
    "is_classical_encounter",
}


def _task_result_rows(
    db: SASession,
    user: User,
    filters: EncounterExportFilters,
) -> list[dict[str, Any]]:
    lab_unit_ids = _scoped_lab_unit_ids(db, user, filters)
    if not lab_unit_ids:
        return []

    params: dict[str, Any] = {
        "lab_unit_ids": lab_unit_ids,
        "hospital_id": filters.hospital_id,
        "lab_unit_id": filters.lab_unit_id,
        "capture_date": filters.capture_date,
        "start_date": filters.start_date,
        "end_date": filters.end_date,
        "project_ids": list(filters.project_ids) or [-1],
        "include_classical": bool(filters.include_classical),
        "has_source_filter": bool(filters.project_ids or filters.include_classical is not None),
    }
    base_column_names = [
        name
        for name in _TASK_RESULT_HEADERS
        if name != "source_view" and name not in _JOINED_TASK_HEADERS
    ]
    select_columns = ", ".join(f"base.{name}" for name in base_column_names)
    sql = sa.text(
        f"""
        WITH disease_task_rows AS (
            SELECT 'dr' AS source_view, *
            FROM mvw_diabetic_retinopathy_grading_pivot
            UNION ALL
            SELECT 'glaucoma' AS source_view, *
            FROM mvw_glaucoma_grading_pivot
            UNION ALL
            SELECT 'amd' AS source_view, *
            FROM mvw_amd_grading_pivot
        ),
        regrade_adj AS (
            SELECT
                g.task_id,
                string_agg(g.id::text, '; ' ORDER BY g.updated_at, g.id) AS regrade_adj_grade_ids,
                string_agg(COALESCE(dg.impression, g.grade_name, ''), '; ' ORDER BY g.updated_at, g.id) AS regrade_adj_grades,
                string_agg(COALESCE(u.username, ''), '; ' ORDER BY g.updated_at, g.id) AS regrade_adj_graders,
                string_agg(COALESCE(g.updated_at::text, ''), '; ' ORDER BY g.updated_at, g.id) AS regrade_adj_grade_times,
                string_agg(COALESCE(g.comment, ''), '; ' ORDER BY g.updated_at, g.id) AS regrade_adj_comments,
                string_agg(COALESCE(g.selected_features_json, ''), '; ' ORDER BY g.updated_at, g.id) AS regrade_adj_features
            FROM grades g
            LEFT JOIN disease_gradings dg ON dg.id = g.disease_grading_id
            LEFT JOIN users u ON u.id = g.grader_user_id
            WHERE g.role_slot = 'regrade_adj'
            GROUP BY g.task_id
        )
        SELECT
            base.source_view,
            {select_columns},
            pe.project_id AS project_id,
            COALESCE(p.code, '') AS project_code,
            COALESCE(p.title, '') AS project_title,
            (pe.project_id IS NULL) AS is_classical_encounter,
            COALESCE(r.regrade_adj_grade_ids, '') AS regrade_adj_grade_ids,
            COALESCE(r.regrade_adj_grades, '') AS regrade_adj_grades,
            COALESCE(r.regrade_adj_graders, '') AS regrade_adj_graders,
            COALESCE(r.regrade_adj_grade_times, '') AS regrade_adj_grade_times,
            COALESCE(r.regrade_adj_comments, '') AS regrade_adj_comments,
            COALESCE(r.regrade_adj_features, '') AS regrade_adj_features
        FROM disease_task_rows base
        JOIN patient_encounters pe ON pe.id = base.patient_encounter_id
        LEFT JOIN projects p ON p.id = pe.project_id
        LEFT JOIN regrade_adj r ON r.task_id = base.task_id
        WHERE pe.lab_unit_id IN :lab_unit_ids
          AND (:hospital_id IS NULL OR pe.lab_unit_id IN (SELECT id FROM lab_units WHERE hospital_id = :hospital_id))
          AND (:lab_unit_id IS NULL OR pe.lab_unit_id = :lab_unit_id)
          AND (:capture_date IS NULL OR pe.capture_date_dt = :capture_date)
          AND (:start_date IS NULL OR pe.capture_date_dt >= :start_date)
          AND (:end_date IS NULL OR pe.capture_date_dt <= :end_date)
          AND (
              :has_source_filter IS FALSE
              OR (:include_classical IS TRUE AND pe.project_id IS NULL)
              OR pe.project_id IN :project_ids
          )
        ORDER BY pe.capture_date_dt DESC NULLS LAST, pe.id DESC, base.image_uuid, base.task_id, base.source_view
        """
    ).bindparams(
        sa.bindparam("lab_unit_ids", expanding=True),
        sa.bindparam("project_ids", expanding=True),
    )
    rows = db.execute(sql, params).mappings().all()
    return [dict(row) for row in rows]


def export_encounter_task_results_xlsx(
    db: SASession,
    user: User,
    filters: EncounterExportFilters,
) -> bytes:
    """Build data-only encounter export with task results and full OCR identifiers."""

    wb = Workbook(write_only=True)
    task_ws = wb.create_sheet("Image Task Results")
    _write_sheet(task_ws, _TASK_RESULT_HEADERS, _task_result_rows(db, user, filters))

    encounters = _load_encounters(db, user, filters)
    max_dr = max((len(encounter.dr_reports) for encounter in encounters), default=0)
    max_glaucoma = max((len(encounter.glaucoma_results_cleaned) for encounter in encounters), default=0)
    ocr_headers = [
        "encounter_id",
        "encounter_uuid",
        "patient_name",
        "patient_id",
        "project_id",
        "project_code",
        "project_title",
        "is_classical_encounter",
        "capture_date_text",
        "capture_date",
        "hospital_id",
        "hospital_name",
        "lab_unit_id",
        "lab_unit_name",
        "zip_file_id",
        "zip_filename",
        "encounter_verified_status",
        "encounter_verified_by",
        "encounter_verified_at",
        "dr_verified_status",
        "dr_verified_by",
        "dr_verified_at",
        "glaucoma_verified_status",
        "glaucoma_verified_by",
        "glaucoma_verified_at",
        "is_set_based",
        "classical_image_count",
        "encounter_set_image_count",
        "image_count",
        "latest_dr_result",
        "latest_dr_qualitative_result",
        "latest_dr_report_file_name",
        "latest_glaucoma_result",
        "latest_glaucoma_qualitative_result",
        "latest_glaucoma_vcdr_right",
        "latest_glaucoma_vcdr_left",
        "latest_glaucoma_report_uuid",
        "latest_glaucoma_report_file_name",
        "dr_ocr_count",
        "glaucoma_ocr_count",
        "remarks",
        "metadata_json",
    ]
    for index in range(max_dr):
        prefix = f"dr_ocr_{index + 1}"
        ocr_headers.extend(
            [
                f"{prefix}_report_id",
                f"{prefix}_report_uuid",
                f"{prefix}_result",
                f"{prefix}_qualitative_result",
                f"{prefix}_report_file_name",
            ]
        )
    for index in range(max_glaucoma):
        prefix = f"glaucoma_ocr_{index + 1}"
        ocr_headers.extend(
            [
                f"{prefix}_cleaned_id",
                f"{prefix}_report_uuid",
                f"{prefix}_result",
                f"{prefix}_qualitative_result",
                f"{prefix}_vcdr_right",
                f"{prefix}_vcdr_left",
                f"{prefix}_original_vcdr_right",
                f"{prefix}_original_vcdr_left",
                f"{prefix}_report_file_name",
                f"{prefix}_created_at",
                f"{prefix}_updated_at",
            ]
        )
    ocr_rows = []
    for encounter in encounters:
        row = _encounter_row(encounter)
        _flatten_ocr_columns(encounter, row, max_dr, max_glaucoma)
        ocr_rows.append(row)
    ocr_ws = wb.create_sheet("Encounter OCR Data")
    _write_sheet(ocr_ws, ocr_headers, ocr_rows)
    return _workbook_bytes(wb)
