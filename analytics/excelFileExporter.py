"""
Excel File Exporter Utility

This utility provides functions to export analytics data as Microsoft XLSX files.
"""

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_encounter_summary_to_xlsx(encounter_data: Dict[str, Any]) -> bytes:
    """
    Export encounter summary data to an XLSX file.
    
    Args:
        encounter_data: Dictionary containing encounter summary data
        
    Returns:
        bytes: XLSX file content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Encounter {encounter_data.get('encounter_id', 'Unknown')}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Add metadata
    ws['A1'] = 'Encounter Summary Report'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    row = 3
    encounter_info = [
        ('Encounter ID', encounter_data.get('encounter_id')),
        ('Patient ID', encounter_data.get('encounter_patient_id')),
        ('Capture Date', encounter_data.get('encounter_capture_date')),
        ('Lab Unit', encounter_data.get('lab_unit_name')),
        ('Hospital', encounter_data.get('hospital_name')),
        ('Total Images', len(encounter_data.get('image_uuids', []))),
        ('Verified Status', encounter_data.get('encounter_verified_status')),
    ]
    
    for key, value in encounter_info:
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=str(value) if value is not None else "N/A")
        row += 1
    
    # Add gap before images section
    row += 1
    
    # Add images with tasks
    ws.cell(row=row, column=1, value="Images and Tasks").font = Font(bold=True, size=14)
    row += 1
    
    # Headers for images with tasks
    headers = ['Image ID', 'UUID', 'Filename', 'Eye Side', 'Task ID', 'Disease', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    # Add image and task data
    images_with_tasks = encounter_data.get('images_with_tasks', [])
    for img_data in images_with_tasks:
        image_id = img_data.get('id')
        uuid = img_data.get('uuid')
        filename = img_data.get('filename')
        eye_side = img_data.get('eye_side')
        
        tasks = img_data.get('tasks', [])
        if tasks:
            for task in tasks:
                ws.cell(row=row, column=1, value=image_id)
                ws.cell(row=row, column=2, value=uuid)
                ws.cell(row=row, column=3, value=filename)
                ws.cell(row=row, column=4, value=eye_side)
                ws.cell(row=row, column=5, value=task.get('id'))
                ws.cell(row=row, column=6, value=task.get('disease'))
                ws.cell(row=row, column=7, value=task.get('status'))
                
                # Apply border to all cells
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
                
                row += 1
        else:
            # Add row with image data but no tasks
            ws.cell(row=row, column=1, value=image_id)
            ws.cell(row=row, column=2, value=uuid)
            ws.cell(row=row, column=3, value=filename)
            ws.cell(row=row, column=4, value=eye_side)
            
            # Apply border to all cells
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
            
            row += 1
    
    # Add glaucoma results cleaned
    row += 1
    ws.cell(row=row, column=1, value="Glaucoma Results Cleaned").font = Font(bold=True, size=14)
    row += 1
    
    gl_headers = ['ID', 'Result', 'Qualitative Result', 'VCDR Right', 'VCDR Left', 'UUID']
    for col, header in enumerate(gl_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    glaucoma_results = encounter_data.get('glaucoma_results_cleaned', [])
    for result in glaucoma_results:
        ws.cell(row=row, column=1, value=result.get('id'))
        ws.cell(row=row, column=2, value=result.get('result'))
        ws.cell(row=row, column=3, value=result.get('qualitative_result'))
        ws.cell(row=row, column=4, value=result.get('vcdr_right_num'))
        ws.cell(row=row, column=5, value=result.get('vcdr_left_num'))
        ws.cell(row=row, column=6, value=result.get('uuid'))
        
        # Apply border to all cells
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    # Add DR reports
    row += 1
    ws.cell(row=row, column=1, value="Diabetic Retinopathy Reports").font = Font(bold=True, size=14)
    row += 1
    
    dr_headers = ['ID', 'Result', 'Qualitative Result', 'UUID']
    for col, header in enumerate(dr_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    dr_reports = encounter_data.get('diabetic_retinopathy_reports', [])
    for report in dr_reports:
        ws.cell(row=row, column=1, value=report.get('id'))
        ws.cell(row=row, column=2, value=report.get('result'))
        ws.cell(row=row, column=3, value=report.get('qualitative_result'))
        ws.cell(row=row, column=4, value=report.get('uuid'))
        
        # Apply border to all cells
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create a BytesIO buffer to save the workbook
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def export_tasks_summary_to_xlsx(tasks_data: List[Dict[str, Any]]) -> bytes:
    """
    Export tasks summary data to an XLSX file.
    
    Args:
        tasks_data: List of dictionaries containing task summary data
        
    Returns:
        bytes: XLSX file content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks Summary"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Add title
    ws['A1'] = 'Tasks Summary Report'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    row = 3
    
    # Add headers
    headers = ['Task ID', 'Status', 'Disease', 'Created At', 'Updated At']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    # Add task data
    for task in tasks_data:
        ws.cell(row=row, column=1, value=task.get('id'))
        ws.cell(row=row, column=2, value=task.get('status'))
        ws.cell(row=row, column=3, value=task.get('disease'))
        ws.cell(row=row, column=4, value=str(task.get('created_at', '')))
        ws.cell(row=row, column=5, value=str(task.get('updated_at', '')))
        
        # Apply border to all cells
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    # Add grades section if there are grades in any task
    has_grades = any(task.get('grades') for task in tasks_data)
    if has_grades:
        row += 1
        ws.cell(row=row, column=1, value="Grades").font = Font(bold=True, size=14)
        row += 1
        
        grade_headers = ['Task ID', 'Grade ID', 'Role', 'Impression', 'Comment', 'Grader', 'Created At']
        for col, header in enumerate(grade_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        row += 1
        
        for task in tasks_data:
            for grade in task.get('grades', []):
                ws.cell(row=row, column=1, value=task.get('id'))
                ws.cell(row=row, column=2, value=grade.get('id'))
                ws.cell(row=row, column=3, value=grade.get('role_slot'))
                ws.cell(row=row, column=4, value=grade.get('impression'))
                ws.cell(row=row, column=5, value=grade.get('comment'))
                ws.cell(row=row, column=6, value=grade.get('grader_username'))
                ws.cell(row=row, column=7, value=str(grade.get('created_at', '')))
                
                # Apply border to all cells
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
                
                row += 1
    
    # Add consensus section if any task has consensus
    has_consensus = any(task.get('consensus') for task in tasks_data)
    if has_consensus:
        row += 1
        ws.cell(row=row, column=1, value="Consensus").font = Font(bold=True, size=14)
        row += 1
        
        consensus_headers = ['Task ID', 'Consensus ID', 'Method', 'Final Impression', 'Decided At', 'Decided By']
        for col, header in enumerate(consensus_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        row += 1
        
        for task in tasks_data:
            consensus = task.get('consensus')
            if consensus:
                ws.cell(row=row, column=1, value=task.get('id'))
                ws.cell(row=row, column=2, value=consensus.get('id'))
                ws.cell(row=row, column=3, value=consensus.get('method'))
                ws.cell(row=row, column=4, value=consensus.get('final_impression'))
                ws.cell(row=row, column=5, value=str(consensus.get('decided_at', '')))
                ws.cell(row=row, column=6, value=consensus.get('decided_by_username'))
                
                # Apply border to all cells
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border
                
                row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create a BytesIO buffer to save the workbook
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def export_image_tasks_to_xlsx(image_uuid: str, image_tasks_data: Dict[str, Any]) -> bytes:
    """
    Export image tasks data to an XLSX file.
    
    Args:
        image_uuid: UUID of the image
        image_tasks_data: Dictionary containing image and tasks data
        
    Returns:
        bytes: XLSX file content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Image Tasks - {image_uuid[:8] if image_uuid else 'Unknown'}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Add title
    ws['A1'] = f'Image Tasks Report - {image_uuid}'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    row = 3
    
    # Add image info
    image_info = image_tasks_data.get('image_info', {})
    ws.cell(row=row, column=1, value="Image Information").font = Font(bold=True, size=14)
    row += 1
    
    img_info_headers = ['Field', 'Value']
    for col, header in enumerate(img_info_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    for key, value in image_info.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=str(value) if value is not None else "N/A")
        
        # Apply border to all cells
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    # Add tasks
    row += 1
    ws.cell(row=row, column=1, value="Associated Tasks").font = Font(bold=True, size=14)
    row += 1
    
    task_headers = ['Task ID', 'Status', 'Disease', 'Created At', 'Updated At']
    for col, header in enumerate(task_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    tasks = image_tasks_data.get('tasks', [])
    for task in tasks:
        ws.cell(row=row, column=1, value=task.get('id'))
        ws.cell(row=row, column=2, value=task.get('status'))
        ws.cell(row=row, column=3, value=task.get('disease'))
        ws.cell(row=row, column=4, value=str(task.get('created_at', '')))
        ws.cell(row=row, column=5, value=str(task.get('updated_at', '')))
        
        # Apply border to all cells
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    # Add grades if any tasks have them
    has_grades = any(task.get('grades') for task in tasks)
    if has_grades:
        row += 1
        ws.cell(row=row, column=1, value="Grades").font = Font(bold=True, size=14)
        row += 1
        
        grade_headers = ['Task ID', 'Grade ID', 'Role', 'Impression', 'Comment', 'Grader', 'Created At']
        for col, header in enumerate(grade_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        row += 1
        
        for task in tasks:
            for grade in task.get('grades', []):
                ws.cell(row=row, column=1, value=task.get('id'))
                ws.cell(row=row, column=2, value=grade.get('id'))
                ws.cell(row=row, column=3, value=grade.get('role_slot'))
                ws.cell(row=row, column=4, value=grade.get('impression'))
                ws.cell(row=row, column=5, value=grade.get('comment'))
                ws.cell(row=row, column=6, value=grade.get('grader_username'))
                ws.cell(row=row, column=7, value=str(grade.get('created_at', '')))
                
                # Apply border to all cells
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
                
                row += 1
    
    # Add consensus if any tasks have it
    has_consensus = any(task.get('consensus') for task in tasks)
    if has_consensus:
        row += 1
        ws.cell(row=row, column=1, value="Consensus").font = Font(bold=True, size=14)
        row += 1
        
        consensus_headers = ['Task ID', 'Consensus ID', 'Method', 'Final Impression', 'Decided At', 'Decided By']
        for col, header in enumerate(consensus_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        row += 1
        
        for task in tasks:
            consensus = task.get('consensus')
            if consensus:
                ws.cell(row=row, column=1, value=task.get('id'))
                ws.cell(row=row, column=2, value=consensus.get('id'))
                ws.cell(row=row, column=3, value=consensus.get('method'))
                ws.cell(row=row, column=4, value=consensus.get('final_impression'))
                ws.cell(row=row, column=5, value=str(consensus.get('decided_at', '')))
                ws.cell(row=row, column=6, value=consensus.get('decided_by_username'))
                
                # Apply border to all cells
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border
                
                row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create a BytesIO buffer to save the workbook
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def export_encounters_summary_list_to_xlsx(encounters_data: List[Dict[str, Any]]) -> bytes:
    """
    Export encounters summary list to an XLSX file.
    
    Args:
        encounters_data: List of dictionaries containing encounter summary data
        
    Returns:
        bytes: XLSX file content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Encounters Summary"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Add title
    ws['A1'] = 'Encounters Summary Report'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    row = 3
    
    # Add headers
    headers = ['ID', 'Name', 'Patient ID', 'Capture Date', 'Image Count', 'Task Count', 'Completed Tasks', 'Lab Unit', 'DR Status', 'Glaucoma Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    # Add encounters data
    for encounter in encounters_data:
        ws.cell(row=row, column=1, value=encounter.get('id'))
        ws.cell(row=row, column=2, value=encounter.get('name'))
        ws.cell(row=row, column=3, value=encounter.get('patient_id'))
        ws.cell(row=row, column=4, value=encounter.get('capture_date'))
        ws.cell(row=row, column=5, value=encounter.get('image_count', 0))
        ws.cell(row=row, column=6, value=encounter.get('task_count', 0))
        ws.cell(row=row, column=7, value=encounter.get('completed_task_count', 0))
        ws.cell(row=row, column=8, value=encounter.get('lab_unit_name'))
        ws.cell(row=row, column=9, value=encounter.get('dr_verified_status'))
        ws.cell(row=row, column=10, value=encounter.get('glaucoma_verified_status'))
        
        # Apply border to all cells
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create a BytesIO buffer to save the workbook
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()