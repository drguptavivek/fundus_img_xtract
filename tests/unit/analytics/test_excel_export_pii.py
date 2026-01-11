"""
Unit tests for Excel Export Pipeline PII Sanitization.

Test IDs from PII_Exposure_Control_Policy.md:
- PII-EXP-001: Excel exports mask patient_id for cross-hospital data
- PII-EXP-004: Export functions accept mask_pii parameter

Bead: 5K (fundus_img_xtract-f6n)

TDD: Tests written FIRST before implementation.
These tests will FAIL until PII masking is implemented.
"""

import pytest
from unittest.mock import Mock, MagicMock, patch
from analytics.excelFileExporter import (
    export_encounter_summary_to_xlsx,
    export_encounters_summary_list_to_xlsx,
    export_encounter_files_to_xlsx
)
import pandas as pd
from openpyxl import load_workbook
import io


class TestEncounterSummaryExport:
    """Tests for export_encounter_summary_to_xlsx PII masking."""
    
    def test_export_encounter_summary_masks_patient_id_when_flag_true(self):
        """
        CRITICAL TEST: export_encounter_summary_to_xlsx must mask patient_id when mask_pii=True.
        
        This test will FAIL until we add mask_pii parameter and implement masking.
        """
        encounter_data = {
            'encounter_id': 123,
            'encounter_patient_id': '12345678',
            'encounter_capture_date': '2024-01-15',
            'lab_unit_name': 'Lab 1',
            'hospital_name': 'Hospital A',
            'image_uuids': ['uuid1', 'uuid2'],
            'encounter_verified_status': 'verified',
            'images_with_tasks': [],
            'glaucoma_results_cleaned': [],
            'diabetic_retinopathy_reports': []
        }
        
        # Export with PII masking enabled
        xlsx_bytes = export_encounter_summary_to_xlsx(encounter_data, mask_pii=True)
        
        # Load the workbook to verify content
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        
        # Find the Patient ID row (should be around row 4-5)
        patient_id_found = False
        for row in ws.iter_rows(min_row=3, max_row=10, values_only=True):
            if row[0] == 'Patient ID':
                patient_id_value = row[1]
                patient_id_found = True
                # Should be masked format: P****678
                assert isinstance(patient_id_value, str), f"Patient ID should be string, got {type(patient_id_value)}"
                assert patient_id_value.startswith('P****'), \
                    f"Patient ID must be masked when mask_pii=True, got: {patient_id_value}"
                assert patient_id_value == 'P****678', \
                    f"Expected 'P****678', got: {patient_id_value}"
                break
        
        assert patient_id_found, "Patient ID row not found in export"
    
    def test_export_encounter_summary_shows_full_patient_id_when_flag_false(self):
        """
        When mask_pii=False, patient_id should be visible (for same-hospital access).
        """
        encounter_data = {
            'encounter_id': 123,
            'encounter_patient_id': '12345678',
            'encounter_capture_date': '2024-01-15',
            'lab_unit_name': 'Lab 1',
            'hospital_name': 'Hospital A',
            'image_uuids': ['uuid1'],
            'encounter_verified_status': 'verified',
            'images_with_tasks': [],
            'glaucoma_results_cleaned': [],
            'diabetic_retinopathy_reports': []
        }
        
        # Export without PII masking
        xlsx_bytes = export_encounter_summary_to_xlsx(encounter_data, mask_pii=False)
        
        # Load the workbook to verify content
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        
        # Find the Patient ID row
        for row in ws.iter_rows(min_row=3, max_row=10, values_only=True):
            if row[0] == 'Patient ID':
                patient_id_value = row[1]
                # Should be full patient_id
                assert patient_id_value == '12345678', \
                    f"Expected full patient_id '12345678', got: {patient_id_value}"
                break


class TestEncountersSummaryListExport:
    """Tests for export_encounters_summary_list_to_xlsx PII masking."""
    
    def test_export_encounters_list_masks_patient_ids_when_flag_true(self):
        """
        CRITICAL TEST: export_encounters_summary_list_to_xlsx must mask all patient_ids when mask_pii=True.
        
        This test will FAIL until we add mask_pii parameter and implement masking.
        """
        encounters_data = [
            {
                'id': 1,
                'name': 'Encounter 1',
                'patient_id': '12345678',
                'capture_date': '2024-01-15',
                'image_count': 5,
                'task_count': 3,
                'completed_task_count': 2,
                'lab_unit_name': 'Lab 1',
                'dr_verified_status': 'verified',
                'glaucoma_verified_status': 'pending'
            },
            {
                'id': 2,
                'name': 'Encounter 2',
                'patient_id': '87654321',
                'capture_date': '2024-01-16',
                'image_count': 3,
                'task_count': 2,
                'completed_task_count': 1,
                'lab_unit_name': 'Lab 2',
                'dr_verified_status': 'pending',
                'glaucoma_verified_status': 'verified'
            }
        ]
        
        # Export with PII masking enabled
        xlsx_bytes = export_encounters_summary_list_to_xlsx(encounters_data, mask_pii=True)
        
        # Load the workbook to verify content
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        
        # Headers are in row 3, data starts at row 4
        # Patient ID is column 3
        patient_ids = []
        for row in ws.iter_rows(min_row=4, max_row=5, min_col=3, max_col=3, values_only=True):
            patient_ids.append(row[0])
        
        # Both patient_ids should be masked
        assert len(patient_ids) == 2
        assert patient_ids[0] == 'P****678', f"Expected 'P****678', got: {patient_ids[0]}"
        assert patient_ids[1] == 'P****321', f"Expected 'P****321', got: {patient_ids[1]}"
    
    def test_export_encounters_list_shows_full_patient_ids_when_flag_false(self):
        """
        When mask_pii=False, all patient_ids should be visible.
        """
        encounters_data = [
            {
                'id': 1,
                'name': 'Encounter 1',
                'patient_id': '12345678',
                'capture_date': '2024-01-15',
                'image_count': 5,
                'task_count': 3,
                'completed_task_count': 2,
                'lab_unit_name': 'Lab 1',
                'dr_verified_status': 'verified',
                'glaucoma_verified_status': 'pending'
            }
        ]
        
        # Export without PII masking
        xlsx_bytes = export_encounters_summary_list_to_xlsx(encounters_data, mask_pii=False)
        
        # Load the workbook to verify content
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        
        # Get patient_id from row 4, column 3
        patient_id = ws.cell(row=4, column=3).value
        assert patient_id == '12345678', f"Expected full patient_id '12345678', got: {patient_id}"


class TestEncounterFilesDataframeExport:
    """Tests for export_encounter_files_to_xlsx with dataframe containing patient_id."""
    
    def test_export_encounter_files_masks_patient_id_column_when_flag_true(self):
        """
        CRITICAL TEST: export_encounter_files_to_xlsx must mask patient_id column when mask_pii=True.
        
        This test will FAIL until we add mask_pii parameter and implement masking.
        """
        # Create a dataframe with patient_id column
        df = pd.DataFrame({
            'encounter_id': [1, 2, 3],
            'patient_id': ['12345678', '87654321', '11223344'],
            'hospital_id': [1, 1, 2],
            'lab_unit_id': [1, 1, 2],
            'capture_date': ['2024-01-15', '2024-01-16', '2024-01-17']
        })
        
        # Export with PII masking enabled
        xlsx_bytes = export_encounter_files_to_xlsx(df, mask_pii=True)
        
        # Load the workbook to verify content
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        
        # Headers are in row 4, data starts at row 5
        # Find patient_id column index
        patient_id_col = None
        for col_idx, cell in enumerate(ws[4], start=1):
            if cell.value and 'Patient' in str(cell.value) and 'Id' in str(cell.value):
                patient_id_col = col_idx
                break
        
        assert patient_id_col is not None, "Patient ID column not found in export"
        
        # Check all patient_id values are masked
        patient_ids = []
        for row_idx in range(5, 8):  # Rows 5, 6, 7 (3 data rows)
            patient_id = ws.cell(row=row_idx, column=patient_id_col).value
            patient_ids.append(patient_id)
        
        # All should be masked
        assert patient_ids[0] == 'P****678', f"Expected 'P****678', got: {patient_ids[0]}"
        assert patient_ids[1] == 'P****321', f"Expected 'P****321', got: {patient_ids[1]}"
        assert patient_ids[2] == 'P****344', f"Expected 'P****344', got: {patient_ids[2]}"
    
    def test_export_encounter_files_shows_full_patient_ids_when_flag_false(self):
        """
        When mask_pii=False, patient_id column should show full values.
        """
        df = pd.DataFrame({
            'encounter_id': [1],
            'patient_id': ['12345678'],
            'hospital_id': [1],
            'capture_date': ['2024-01-15']
        })
        
        # Export without PII masking
        xlsx_bytes = export_encounter_files_to_xlsx(df, mask_pii=False)
        
        # Load the workbook to verify content
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        
        # Find patient_id column
        patient_id_col = None
        for col_idx, cell in enumerate(ws[4], start=1):
            if cell.value and 'Patient' in str(cell.value) and 'Id' in str(cell.value):
                patient_id_col = col_idx
                break
        
        # Get patient_id value
        patient_id = ws.cell(row=5, column=patient_id_col).value
        assert patient_id == '12345678', f"Expected full patient_id '12345678', got: {patient_id}"


class TestDefaultBehavior:
    """Tests for default behavior when mask_pii parameter is not provided."""
    
    def test_export_functions_default_to_no_masking(self):
        """
        By default (mask_pii not specified), exports should NOT mask PII.
        
        This maintains backward compatibility with existing code.
        """
        encounter_data = {
            'encounter_id': 123,
            'encounter_patient_id': '12345678',
            'encounter_capture_date': '2024-01-15',
            'lab_unit_name': 'Lab 1',
            'hospital_name': 'Hospital A',
            'image_uuids': [],
            'encounter_verified_status': 'verified',
            'images_with_tasks': [],
            'glaucoma_results_cleaned': [],
            'diabetic_retinopathy_reports': []
        }
        
        # Export without specifying mask_pii (should default to False)
        xlsx_bytes = export_encounter_summary_to_xlsx(encounter_data)
        
        # Load and verify - should show full patient_id
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        
        for row in ws.iter_rows(min_row=3, max_row=10, values_only=True):
            if row[0] == 'Patient ID':
                patient_id_value = row[1]
                # Should be full patient_id (backward compatible)
                assert patient_id_value == '12345678', \
                    f"Default behavior should show full patient_id, got: {patient_id_value}"
                break
