from __future__ import annotations

import io
from datetime import date, datetime, timezone
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from encounter_sets import export_service
from encounter_sets.export_service import (
    EncounterSetExportFilters,
    EncounterSetExportValidationError,
    export_encounter_sets_xlsx,
    parse_export_month,
)
from models import AMDReport, DiabeticRetinopathyReport, GlaucomaReport, GlaucomaResultsCleaned


def test_parse_export_month_uses_exclusive_next_month_boundary():
    assert parse_export_month("2026-07") == (date(2026, 7, 1), date(2026, 8, 1))
    assert parse_export_month("2026-12") == (date(2026, 12, 1), date(2027, 1, 1))


def test_parse_export_month_rejects_invalid_values():
    with pytest.raises(EncounterSetExportValidationError, match="YYYY-MM"):
        parse_export_month("2026-7")


def test_export_contains_emr_fields_timezone_pdf_flags_and_all_ocr_columns(monkeypatch):
    dr_one = DiabeticRetinopathyReport(
        id=11,
        patient_encounter_id=3477,
        uuid="dr-one",
        result="No DR",
        qualitative_result="Clear",
        report_file_name="ai.pdf",
    )
    dr_two = DiabeticRetinopathyReport(
        id=12,
        patient_encounter_id=3477,
        uuid="dr-two",
        result="Mild DR",
        qualitative_result="Review",
        report_file_name="repeat.pdf",
    )
    glaucoma = GlaucomaReport(
        id=21,
        patient_encounter_id=3477,
        uuid="gl-one",
        vcdr_right="0.70",
        vcdr_left="0.60",
        result="Suspect",
        qualitative_result="Refer",
        report_file_name="ai.pdf",
    )
    cleaned = GlaucomaResultsCleaned(
        id=22,
        glaucoma_report_id=21,
        patient_encounter_id=3477,
        vcdr_right_num=0.7,
        vcdr_left_num=0.6,
        original_vcdr_right="VCDR 0.70",
        original_vcdr_left="VCDR 0.60",
        result="Suspect",
        qualitative_result="Refer",
        report_uuid="gl-one",
        report_file_name="ai.pdf",
        created_at=datetime(2026, 7, 31, 10, tzinfo=timezone.utc),
        updated_at=datetime(2026, 7, 31, 11, tzinfo=timezone.utc),
    )
    amd = AMDReport(
        id=31,
        patient_encounter_id=3477,
        uuid="amd-one",
        result="AMD detected",
        qualitative_result="Refer",
        report_file_name="ai.pdf",
    )
    attachment = SimpleNamespace(
        asset_kind="pdf",
        mime_type="application/pdf",
        metadata_json={
            "ocr": {
                "dr_report": {"diabetic_retinopathy_report_id": 11},
                "glaucoma_report": {"glaucoma_report_id": 21},
                "amd_report": {"amd_report_id": 31},
            }
        },
    )
    encounter = SimpleNamespace(
        id=3477,
        patient_id="UHID-44",
        name="Fallback Name",
        capture_date="2026-07-31",
        capture_date_dt=date(2026, 7, 31),
        metadata_json={
            "patient": {
                "hospital_UHID": "UHID-44",
                "patient_name": "Patient Name",
                "patient_age_yrs": 64,
                "sex": "F",
                "remidio_site_custom_identifier": "SITE-METADATA",
            },
            "encounter": {"capture_datetime": "2026-07-31T18:45:30Z"},
        },
        encounter_set_images=[SimpleNamespace(id=1), SimpleNamespace(id=2)],
        encounter_set_attachments=[attachment],
        dr_reports=[dr_one, dr_two],
        glaucoma_reports=[glaucoma],
        glaucoma_results_cleaned=[cleaned],
        amd_reports=[amd],
    )
    exam = SimpleNamespace(
        patient_encounter_id=3477,
        site_custom_identifier="SITE-DELHI",
        exam_date=None,
    )
    monkeypatch.setattr(
        export_service, "_load_encounters", lambda *_args, **_kwargs: [encounter]
    )
    monkeypatch.setattr(export_service, "_load_remidio_exams", lambda *_args: {3477: exam})

    content = export_encounter_sets_xlsx(
        object(),
        user=object(),
        filters=EncounterSetExportFilters(project_id=3, month="2026-07"),
        timezone_name="Asia/Kolkata",
        include_identifiers=True,
    )
    sheet = load_workbook(io.BytesIO(content), read_only=True).active
    headers = list(next(sheet.iter_rows(values_only=True)))
    values = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    row = dict(zip(headers, values, strict=True))

    assert row["EncounterID"] == 3477
    assert row["Type"] == "encounterSet"
    assert row["hospital_UHID"] == "UHID-44"
    assert row["patient_name"] == "Patient Name"
    assert row["patient_age_yrs"] == 64
    assert row["sex"] == "F"
    assert row["remidio_site_custom_identifier"] == "SITE-DELHI"
    assert row["capture_date"] == "2026-08-01"
    assert row["capture_time"] == "00:15:30"
    assert row["clinical_image_count"] == 2
    assert row["has_DR_PDF"] is True
    assert row["has_glaucoma_PDF"] is True
    assert row["has_AMD_PDF"] is True
    assert row["dr_ocr_1_result"] == "No DR"
    assert row["dr_ocr_2_result"] == "Mild DR"
    assert row["glaucoma_ocr_1_vcdr_right"] == "0.70"
    assert row["glaucoma_cleaned_ocr_1_vcdr_right_num"] == 0.7
    assert row["amd_ocr_1_result"] == "AMD detected"
    assert "dr_ocr_1_patient_encounter_id" in headers
    assert "glaucoma_cleaned_ocr_1_updated_at" in headers


def test_pdf_flags_require_disease_evidence_on_a_pdf_attachment():
    attachment = SimpleNamespace(
        asset_kind="document",
        mime_type="text/plain",
        metadata_json={"ocr": {"dr_report": {"result": "No DR"}}},
    )
    assert export_service._has_disease_pdf([attachment], "dr") is False


def test_non_remidio_flat_metadata_fills_demographics_and_capture_time():
    encounter = SimpleNamespace(
        id=4001,
        patient_id="IITK-9",
        name="MRNIITK-9",
        capture_date="2026-07-15",
        capture_date_dt=date(2026, 7, 15),
        metadata_json={
            "source_kind": "iitk_zip",
            "age": 51,
            "gender": "M",
            "started_at": "2026-07-15T08:30:00Z",
            "site": "IITK",
        },
        encounter_set_images=[SimpleNamespace(id=1)],
        encounter_set_attachments=[],
        dr_reports=[],
        glaucoma_reports=[],
        glaucoma_results_cleaned=[],
        amd_reports=[],
    )

    row = export_service._encounter_row(
        encounter,
        None,
        export_service._target_timezone("Asia/Kolkata"),
        {prefix: 0 for prefix, _model, _relationship in export_service._OCR_MODELS},
        include_identifiers=True,
    )

    assert row["hospital_UHID"] == "IITK-9"
    assert row["patient_age_yrs"] == 51
    assert row["sex"] == "M"
    assert row["capture_date"] == "2026-07-15"
    assert row["capture_time"] == "14:00:00"
    assert row["remidio_site_custom_identifier"] is None


@pytest.mark.parametrize("value", ["=1+1", "+cmd", "-2+3", "@SUM(A1:A2)"])
def test_xlsx_value_neutralizes_formula_prefixes(value):
    assert export_service._xlsx_value(value) == f"'{value}"
